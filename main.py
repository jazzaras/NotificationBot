from datetime import datetime
import pandas	
import xlsxwriter
from openpyxl import load_workbook
import os
import telebot



BASE_DIR = os.path.dirname(os.path.abspath(__file__))

path = os.path.join(BASE_DIR, 'data.xlsx')

filename = 'data.xlsx'

hour = ""
minute = ""
wantedTime = f"{hour}:{minute}"

bot = telebot.TeleBot(BOT_TOKEN)




while True:

	df = pandas.read_excel(filename, sheet_name='main')
	wb = load_workbook(filename)
	ws = wb.worksheets[0]


	# Getting information from excel sheet
	ids = df['Chatid'].tolist()
	day = df['Day'].tolist()
	time = df['Time'].tolist()	
	done = df['Done'].tolist()

	# Getting times and dates
	current = datetime.now()
	dayName = current.strftime('%a').lower()
	currentTime = current.strftime("%H:%M")






	# loop for x times --> x is the number of ids
	for i in range(len(ids)):

		userdays = day[i]
		userdays = userdays.replace("[", "")
		userdays = userdays.replace("]", "")
		userdays = userdays.replace("'", "")
		userdays = userdays.split(",")
		userdays = [Day.strip() for Day in userdays]

		usertimes = time[i]
		usertimes = usertimes.replace("[", "")
		usertimes = usertimes.replace("]", "")
		usertimes = usertimes.replace("'", "")
		usertimes = usertimes.split(",")
		usertimes = [(time.strip()).lower() for time in usertimes]


		for x in range(len(userdays)):
			if dayName == userdays[x]:
				if currentTime == usertimes[x]:
					if not done[i]:
						# Send Noti0fication

						bot.send_message(ids[i], "you have a class now")

						# Set Done to True --> so it will not go in the if statment for a whole minute
						cell = "D" + str(i+2)
						ws[cell] = True
						wb.save(filename)

						print("sent")


					
	# Reset ALL Done boolean variables in excel				
	
	if currentTime == "00:00":

		for i in range(len(ids)):

			cell = "D" + str(i+2)

			ws[cell] = False

			wb.save(filename)			





# print("The current date and time is", currentTime, type(currentTime))